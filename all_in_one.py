from openpyxl import workbook
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import time
import re
import timeit


def change_numbers_for_date(text):
    text = text.replace('۳۱', '۳۰')
    text = text.replace('۰', '0')
    text = text.replace('۱', '1')
    text = text.replace('۲', '2')
    text = text.replace('۳', '3')
    text = text.replace('۴', '4')
    text = text.replace('۵', '5')
    text = text.replace('۶', '6')
    text = text.replace('۷', '7')
    text = text.replace('۸', '8')
    text = text.replace('۹', '9')
    text = text.replace('/', '-')
    return text


def change_numbers(text):
    text = text.replace('۰', '0')
    text = text.replace('۱', '1')
    text = text.replace('۲', '2')
    text = text.replace('۳', '3')
    text = text.replace('۴', '4')
    text = text.replace('۵', '5')
    text = text.replace('۶', '6')
    text = text.replace('۷', '7')
    text = text.replace('۸', '8')
    text = text.replace('۹', '9')
    return text


def rep_char(text, header):
    if header is False:
        text = text.replace(')', '')
        text = text.replace('(', '-')
    text = text.replace(',', '')
    text = text.replace('\n', '')
    text = text.replace('\u200c', ' ')
    text = text.replace('-زیان', '(زیان)')
    text = text.replace('-زيان', '(زيان)')
    text = text.replace('-خروج', '(خروج)')
    text = text.replace('-کاهش', '(کاهش)')
    text = text.replace('-(کسر)', 'کسر')
    text = text.replace('\xa0', '')
    text = text.replace('\u200f', '')
    text = change_numbers(text)
    return text


def str_to_int_or_float(value):
    if isinstance(value, bool):
        return value
    try:
        return int(value)
    except ValueError:
        try:
            return float(value)
        except ValueError:
            return value


def get_stock_names(loc, column, from_which_stock, to_which_stock):
    wb = openpyxl.load_workbook(loc)
    sheet = wb.active
    names = list(sheet.columns)[column]
    names_list = []
    i = from_which_stock - 1
    while i <= to_which_stock - 1:
        names_list.append(str(names[i].value))
        i += 1
    return names_list


class cell:
    def __init__(self):
        self.rowspan = 0
        self.colspan = 0
        self.data = ''

    def add_rowspan(self, rowspan):
        self.rowspan = int(rowspan)

    def add_colspan(self, colspan):
        self.colspan = int(colspan)

    def add_data(self, data):
        self.data = data

    def print_span(self):
        print(self.rowspan, self.colspan)

    def print_data(self):
        print(self.data)

    def give_rowspan(self):
        return self.rowspan

    def give_colspan(self):
        return self.colspan

    def give_data(self):
        return self.data


class table:
    @staticmethod
    def find_colspan(data, line):
        try:
            data.add_colspan(line.get('colspan'))
        except:
            data.add_colspan(0)

    @staticmethod
    def find_rowspan(data, line):
        try:
            data.add_rowspan(line.get('rowspan'))
        except:
            data.add_rowspan(0)

    @staticmethod
    def class_attribute(line):
        try:
            class_element = line.get('class')
            if class_element == None:
                return False
            if isinstance(class_element, list):
                cul = ''
                for w in class_element:
                    cul += w
                class_element = cul
            if class_element.find('Hidden') != -1:
                return True
            else:
                return False
        except:
            return None

    @staticmethod
    def hidden_attribute(line):
        try:
            hidden_element = line.get('hidden')
            if hidden_element == None:
                return False
            else:
                return True
        except:
            return False

    @staticmethod
    def style_attribute(line):
        try:
            style_element = line.get('style')
            if style_element == None:
                return False
            if style_element.find('display:none') != -1:
                return True
            else:
                return False
        except:
            return False

    @staticmethod
    def input_tag(line):
        try:
            if line.find('span') != None:
                return False
            input_element = line.find('input')
            if input_element == None:
                return False
            type_element = input_element.get('type')
            if type_element == None:
                return False
            x = input_element['value']
            if type_element == 'text':
                return True
            else:
                return False
        except:
            return False

    @staticmethod
    def get_cols(row, row_type):
        satr = []
        if row_type == 'th':
            header = True
        else:
            header = False
        for col in row.find_all(row_type):
            one_cell = cell()
            table.find_colspan(one_cell, col)
            table.find_rowspan(one_cell, col)
            if table.class_attribute(col):
                continue
            if table.hidden_attribute(col):
                continue
            if table.style_attribute(col):
                continue
            if table.input_tag(col):
                input_element = col.find('input')
                one_cell.add_data(rep_char(input_element['value'], header))
            else:
                one_cell.add_data(rep_char(col.get_text(), header))
            satr.append(one_cell)
        return satr

    @classmethod
    def get_table(cls, table, hole_data):
        class_table = cls()
        for row in table.find_all("tr"):
            if cls.class_attribute(row):
                continue
            if len(row.find_all("th")) != 0:
                satr = class_table.get_cols(row, 'th')
            else:
                satr = class_table.get_cols(row, 'td')
            hole_data.append(satr)
        return hole_data


class codal_table:
    @staticmethod
    def first_info_table(table):
        if str(table).find('PeriodExtraDay') != -1:
            return True
        else:
            return False

    @staticmethod
    def body_table(soup):
        my_table = soup.find_all('table')
        if codal_table.first_info_table(my_table):
            my_table.pop(0)
        t_body = my_table[0].find("tbody")
        hole_data = []
        hole_data = table.get_table(t_body, hole_data)
        return hole_data

    @staticmethod
    def head_body_table(soup):
        my_table = soup.find('table')
        t_head = my_table.find("thead")
        hole_data = []
        hole_data = table.get_table(t_head, hole_data)
        t_body = my_table.find("tbody")
        hole_data = table.get_table(t_body, hole_data)
        return hole_data

    @staticmethod
    def two_body_table(soup):
        t_body = soup.find_all("tbody")
        if codal_table.first_info_table(t_body):
            t_body.pop(0)
        hole_data = []
        hole_data = table.get_table(t_body[0], hole_data)
        hole_data = table.get_table(t_body[1], hole_data)
        return hole_data

    @staticmethod
    def two_table(soup):
        if len(soup.find_all('table')) != 3:
            raise Exception("not a two_table")
        main_table = soup.find_all('table')
        if codal_table.first_info_table(main_table):
            main_table.pop(0)
        t_body = main_table[1].find("tbody")
        hole_data = []
        hole_data = table.get_table(t_body, hole_data)
        t_body = main_table[2].find("tbody")
        hole_data = table.get_table(t_body, hole_data)
        return hole_data

    @classmethod
    def final_table(cls, soup):
        try:
            x = cls.head_body_table(soup)
            print('head_body_table')
            return x
        except:
            try:
                x = cls.two_table(soup)
                print('two_table')
                return x
            except:
                try:
                    x = cls.two_body_table(soup)
                    print('two_body_table')
                    return x
                except:
                    x = cls.body_table(soup)
                    print('body_table')
                    return x
class make_file :
    @staticmethod
    def make_empty_cells(data):
        i = 0
        empty_cell = cell()
        while i < len(data):
            j = 0
            while j < len(data[i]):
                for k in range(1, data[i][j].give_rowspan()):
                    data[i + k].insert(j, empty_cell)
                for k in range(data[i][j].give_colspan() - 1):
                    data[i].insert(j + 1, empty_cell)
                j += 1
            i += 1
        return data

    @staticmethod
    def make_excel(name, stock_id, data, attribute):
        wb = Workbook()
        wb.save('codal'+ '\\' + stock_id + '\\' + name + '.xlsx')

        x = []
        x.append(name)
        x += attribute
        ws = wb.active
        ws.title = "Page 1"
        ws.append(x)
        data = make_file.make_empty_cells(data)
        for i in range(len(data)):
            temp = []
            for j in range(len(data[i])):
                try:
                    temp.append(str_to_int_or_float(data[i][j].give_data()))
                except:
                    temp.append(data[i][j].give_data())
            ws.append(temp)
        wb.save('codal'+ '\\' + stock_id + '\\' + name + '.xlsx')

    @staticmethod
    def make_folder(id):
        try:
            os.mkdir('codal')
        except:
            pass
        try:
            os.mkdir('codal' + '\\' +str(id))
        except:
            pass

    @staticmethod
    def make_sub_file(stock_id, company, sub_id, company_state_code):
        try:
            wb = openpyxl.load_workbook('codal'+ '\\' + stock_id + '\\' + stock_id + ".xlsx")
            ws = wb.active
        except:
            wb = Workbook()
            wb.save('codal'+ '\\' + stock_id + '\\' + stock_id + ".xlsx")
            ws = wb.active
            ws.title = "Page 1"
            ws.cell(row=1, column=1).value = 'Stock id'
            ws.cell(row=1, column=2).value = 'Main company name'
            ws.cell(row=1, column=3).value = 'Company status'
            ws.cell(row=2, column=1).value = stock_id
            if sub_id is None:
                ws.cell(row=2, column=2).value = str(company)
                ws.cell(row=2, column=3).value = company_state_code
            ws.cell(row=3, column=1).value = 'Sub id'
            ws.cell(row=3, column=2).value = 'Sub company name'
            ws.cell(row=3, column=3).value = 'Company status'
            for i in range(1, 10):
                ws.cell(row=i + 3, column=1).value = '0' + str(i)
            for j in range(10, 100):
                ws.cell(row=j + 3, column=1).value = str(j)

        if sub_id == None and ws.cell(row=2, column=2).value == None:
            ws.cell(row=2, column=2).value = str(company)
            ws.cell(row=2, column=3).value = company_state_code
        if sub_id != None:
            for i in range(1, 100):
                if int(sub_id) == i and ws.cell(row=i + 3, column=2).value == None:
                    ws.cell(row=i + 3, column=2).value = str(company)
                    ws.cell(row=i + 3, column=3).value = company_state_code
        wb.save('codal'+ '\\' + stock_id + '\\' + stock_id + ".xlsx")

    @staticmethod
    def make_error_file(stock_id, stock_type, link):
        try:
            wb = openpyxl.load_workbook('codal'+ '\\' + stock_id + '\\' + stock_id + '_errors' + ".xlsx")
            ws = wb.active
        except:
            wb = Workbook()
            wb.save('codal'+ '\\' + stock_id + '\\' + stock_id + '_errors' + ".xlsx")
            ws = wb.active
            ws.title = "Page 1"
        x = []
        print('Found unreadable table , link :', link)
        x.append(stock_type)
        x.append(link)
        ws.append(x)
        wb.save('codal'+ '\\' + stock_id + '\\' + stock_id + '_errors' + ".xlsx")


class stock_codal:
    second = 0.001
    path = "chromedriver.exe"
    url = "https://codal.ir/"
    codal_loading = 'col-12 loading ng-scope'
    stock_file_names = 'stock_names.xlsx'
    id_column = 0
    name_column = 1
    first_page_url = 'https://codal.ir/ReportList.aspx?search&Symbol=findme&LetterType=-1&Isic=722008&AuditorRef=-1&PageNumber=1&Audited&NotAudited&IsNotAudited=false&Childs&Mains&Publisher=false&CompanyState=0&Category=-1&CompanyType=1&Consolidatable&NotConsolidatable'

    company_id = 'ctl00_txbCompanyName'
    capital_id = 'ctl00_lblListedCapital'
    symbol_id = 'ctl00_txbSymbol'
    unauthorized_capital_id = 'ctl00_txbUnauthorizedCapital'
    ISIC_id = 'ctl00_lblISIC'
    period_id = 'ctl00_lblPeriod'
    period_end_id = 'ctl00_lblPeriodEndToDate'
    date_id = 'ctl00_lblYearEndToDate'
    company_state_id = 'ctl00_lblCompanyState'

    def __init__(self, my_name, my_id):
        self.name = my_name
        self.id = my_id
        self.status = True
        self.all_activities = []
        self.num_activities = 0
        self.all_fiscals = []
        self.num_fiscals = 0

    def get_activity(self, my_activity):
        self.all_activities.append(my_activity)
        self.num_activities += 1

    def get_fiscal(self, my_fiscal):
        self.all_fiscals.append(my_fiscal)
        self.num_fiscals += 1

    def remove_activity(self, index):
        print('removeeeeeeeeeeeeeeeeeeeeeee', self.all_activities.pop(index))
        self.num_activities -= 1

    def remove_fiscal(self, index):
        print('removeeeeeeeeeeeeeeeeeeeeeee', self.all_fiscals.pop(index))
        self.num_fiscals -= 1

    def print_name(self):
        print('** name of stock is', self.name)

    def print_all_activities(self):
        print('** list of activities for stock', self.name, ':')
        for i in range(len(self.all_activities)):
            print('    ', i + 1, self.all_activities[i])

    def print_all_fiscals(self):
        print('** list of fiscals for stock', self.name, ':', )
        for i in range(len(self.all_fiscals)):
            print('    ', i + 1, self.all_fiscals[i])

    def print_num_fiscals(self):
        print('Stock ', self.name, ' have ', self.num_fiscals, 'fiscals')

    def print_num_activities(self):
        print('Stock ', self.name, ' have ', self.num_activities, 'activities')

    def open_first_page(self, browser):
        browser.get(self.first_page_url.replace('findme', self.name))
        time.sleep(15)

    @classmethod
    def get_stock_names(cls, path, id_column, name_column, start, end):
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        names = list(sheet.columns)[name_column]
        ids = list(sheet.columns)[id_column]
        x = []
        for i in range(start, end + 1):
            if names[i - 1].value != None:
                x.append(cls(str(names[i - 1].value), str(ids[i - 1].value)))
        return x

    @staticmethod
    def find_all_li(browser):
        page = browser.page_source
        soup = BeautifulSoup(page, 'lxml')
        my_nav = soup.find_all('nav')[1]
        my_ul = my_nav.find('ul')
        li = my_ul.find_all("li")
        return li

    @staticmethod
    def find_num_pages(li):
        number_of_pages = li[len(li) - 4].find('a')
        number_of_pages = number_of_pages.get_text()
        number_of_pages = change_numbers(number_of_pages)
        number_of_pages = int(number_of_pages)
        return number_of_pages

    @staticmethod
    def find_date(line):
        match = re.search(r'\d{4}-\d{2}-\d{2}', line)
        if match is None:
            match = re.search(r'\d{2}-\d{2}-\d{4}', line)
        if match is None:
            match = re.search(r'\d{2}-\d{2}-\d{2}', line)
        if match is None:
            print("ERROR___there is sth wrong with data:", line)
        return match

    @staticmethod
    def find_period(line):
        if line.find('1') != -1:
            return 1
        elif line.find('2') != -1:
            return 2
        elif line.find('3') != -1:
            return 3
        elif line.find('6') != -1:
            return 6
        elif line.find('9') != -1:
            return 9
        elif line.find('12') != -1:
            return 12
        else:
            return None

    @staticmethod
    def find_link(line):
        line_a = line.find("a")
        link = line_a['href']
        link = 'https://codal.ir' + link
        return link

    @staticmethod
    def find_sub(line, audit, correction):
        if correction:
            y = line.find('(اصلاحیه)')
            x = line[y + 9:]
        else:
            if audit:
                y = line.find('(حسابرسی شده)')
            else:
                y = line.find('(حسابرسی نشده)')
            x = line[y + 14:]
        if x == '':
            return "main"
        else:
            return x

    @staticmethod
    def check_state(state):
        if state.find('فرابورس') != -1:
            return 'Stock morket'
        elif state.find('بورس') != -1:
            return 'OTC market'
        elif state.find('نشده'):
            return 'Not accepted'
        else:
            return 'Other'

    @staticmethod
    def find_month(date):
        if date.find('/01/') != -1:
            return '01'
        elif date.find('/02/') != -1:
            return '02'
        elif date.find('/03/') != -1:
            return '03'
        elif date.find('/04/') != -1:
            return '04'
        elif date.find('/05/') != -1:
            return '05'
        elif date.find('/06/') != -1:
            return '06'
        elif date.find('/07/') != -1:
            return '07'
        elif date.find('/08/') != -1:
            return '08'
        elif date.find('/09/') != -1:
            return '09'
        elif date.find('/10/') != -1:
            return '10'
        elif date.find('/11/') != -1:
            return '11'
        elif date.find('/12/') != -1:
            return '12'

    @staticmethod
    def open_browser(path):
        option = Options()
        option.add_argument("--disable-infobars")
        option.add_argument("start-maximized")
        option.add_argument("--disable-extensions")
        option.add_experimental_option("prefs", {"profile.default_content_setting_values.notifications": 1})
        return webdriver.Chrome(options=option, executable_path=path)

    def open_codal(self, browser):
        while True:
            try:
                browser.get(self.url)
                print("SUCCESS___I could open the main page.")
                time.sleep(1)
                search = browser.find_element_by_id('aSearch')
                search.click()
                print("SUCCESS___I could click on search.")
                time.sleep(1)
                search_space = browser.find_element_by_xpath('//*[@id="collapse-search-1"]/div[2]/div[1]/div/div/a')
                search_space.click()
                time.sleep(1)
                search_box = browser.find_element_by_id('txtSymbol')
                search_box.send_keys(self.name)
                break
            except:
                print("ERROR___I could NOT open the main page.")
                time.sleep(1)

    def get_search_result(self, browser):
        while True:
            try:
                load_element = browser.find_element_by_class_name(self.codal_loading)
                time.sleep(stock_codal.second)
            except:
                search_result = None
                try:
                    search_result = browser.find_element_by_xpath(
                        '//*[@id="ui-select-choices-row-0-0"]/div/div[1]/span')
                except:
                    try:
                        search_result = browser.find_element_by_xpath('//*[@id="ui-select-choices-row-0-0"]/div/div[1]')
                    except:
                        try:
                            search_result = browser.find_element_by_xpath(
                                '//*[@id="ui-select-choices-row-0-0"]/div/div[2]')
                        except:
                            print('cant search stock', self.name)
                return search_result

    def is_it_bad_stock(self, browser, search_result):
        times = 0
        while True:
            try:
                load_element = browser.find_element_by_class_name(self.codal_loading)
                time.sleep(self.second)
            except:
                try:
                    search_result.click()
                    return False
                except:
                    if times == 5:
                        self.status = False
                        return True
                        print(self.name, 'is a bad stock')
                    else:
                        time.sleep(1)
                        times += 1

    def check_get_li(self, browser):
        while True:
            try:
                load_element = browser.find_element_by_class_name(self.codal_loading)
                time.sleep(self.second)
            except:
                try:
                    return self.find_all_li(browser)
                except:
                    time.sleep(self.second)

    def next_page_address(self, browser, stock_url, counting):
        y = stock_url.find('PageNumber')
        stock_url = stock_url[:y + 11] + str(counting + 1) + stock_url[y + 12:]
        while True:
            try:
                load_element = browser.find_element_by_class_name(self.codal_loading)
                time.sleep(self.second)
            except:
                return stock_url

    def accessibility_func(self, browser):
        times = 0
        number_of_pages = 100000
        while True:
            try:
                load_element = browser.find_element_by_class_name(self.codal_loading)
                time.sleep(self.second)
            except:
                try:
                    li = self.find_all_li(browser)
                    number_of_pages = self.find_num_pages(li)
                except:
                    if times == 5:
                        print("Page", self.name, 'Not accessable')
                        self.status = False
                        return False
                    else:
                        time.sleep(1)
                        times += 1
                if number_of_pages <= 5000:
                    print("namad", self.name, "has", number_of_pages, "pages")
                    return True

    def open_report_page(self, url, browser):
        while True:
            try:
                browser.get(url)
                print("success___I could open the link page")
                break
            except:
                print("ERROR___I could NOT open the link page :", url)
                time.sleep(self.second)
        while True:
            try:
                main_page = browser.page_source
                main_soup = BeautifulSoup(main_page, 'lxml')
                main_menu = main_soup.find_all('option')
                print("success___I could make a soup")
                return main_menu, browser
            except:
                print("ERROR___I could not make a soup")
                time.sleep(self.second)

    def click_menu(self, browser, i):
        try:
            menu = browser.find_element_by_xpath('/html/body/form/div[4]/div[3]/select/option[' + str(i) + ']')
            menu.click()
        except:
            try:
                menu = browser.find_element_by_xpath('/html/body/form/div[4]/div[3]/select/option')
                menu.click()
            except:
                try:
                    menu = browser.find_element_by_xpath('/html/body/form/div[3]/div[3]/select/option[' + str(i) + ']')
                    menu.click()
                except:
                    try:
                        menu = browser.find_element_by_xpath('/html/body/form/div[3]/div[3]/select/option')
                        menu.click()
                    except:
                        pass


class activity(stock_codal):
    def __init__(self):
        self.company = None
        self.symbol = None
        self.capital = None
        self.unauthorized_capital = None
        self.period_end = None
        self.date = None
        self.company_state = None
        self.link = None

    def print_activity(self):
        print(self.symbol, self.company, self.period_end \
              , self.capital, self.unauthorized_capital, self.date, self.company_state, self.link)

    def print_list_of_activities(self, list_act):
        for i in list_act:
            print(i.date, i.link)

    def print_link(self):
        print(self.my_link)

    def add_link(self, link):
        self.link = link

    def add_company(self, soup):
        company_name = soup.find(id=self.company_id)
        if company_name is None:
            return
        company_name = company_name.get_text()
        self.company = str(company_name)

    def add_symbol(self, soup):
        symbol_name = soup.find(id=self.symbol_id)
        if symbol_name is None:
            return
        symbol_name = symbol_name.get_text()
        self.symbol = str(symbol_name)

    def add_capital(self, soup):
        capital_name = soup.find(id=self.capital_id)
        if capital_name is None:
            return
        capital_name = capital_name.get_text()
        capital_name = capital_name.replace(',', '')
        self.capital = int(capital_name)

    def add_unauthorized_capital(self, soup):
        unauthorized_capital_name = soup.find(id=self.unauthorized_capital_id)
        if unauthorized_capital_name is None:
            return
        unauthorized_capital_name = unauthorized_capital_name.get_text()
        unauthorized_capital_name = unauthorized_capital_name.replace(',', '')
        self.unauthorized_capital = str(unauthorized_capital_name)

    def add_period_end(self, soup):
        period_end_name = soup.find(id=self.period_end_id)
        if period_end_name is None:
            return
        period_end_name = period_end_name.get_text()
        period_end_name = self.find_month(period_end_name)
        self.period_end = str(period_end_name)

    def add_date(self, soup):
        date_name = soup.find(id=self.date_id)
        if date_name is None:
            return
        date_name = date_name.get_text()
        y = date_name.find('13')
        date_name = date_name[y:y + 4]
        self.date = str(date_name)

    def add_company_state(self, soup):
        company_state_name = soup.find(id=self.company_state_id)
        if company_state_name is None:
            return
        company_state_name = company_state_name.get_text()
        company_state_name = self.check_state(company_state_name)
        self.company_state = str(company_state_name)

    def make_name(self):
        name = str(self.date) + '-' + str(self.period_end)
        return name

    def save_attribute(self):
        att = [self.company, self.symbol, self.capital, self.unauthorized_capital, self.company_state]
        return att

    def sub_id(self):
        x = re.search(r'\d+', self.symbol)
        if x is None:
            return '00'
        elif int(x.group()) < 10:
            return '0' + x.group()
        elif int(x.group()) >= 10:
            return x.group()

    def code_company_state(self):
        if self.company_state == 'Stock morket':
            return '01'
        elif self.company_state == 'OTC market':
            return '02'
        elif self.company_state == 'Other':
            return '03'
        else:
            return 'N/A'

    def check_existence(self):
        if self.company is None:
            return False
        if self.symbol is None:
            return False
        if self.capital is None:
            return False
        return True

    def activity_analysis(self, soup):
        self.add_company(soup)
        self.add_symbol(soup)
        self.add_capital(soup)
        self.add_unauthorized_capital(soup)
        self.add_period_end(soup)
        self.add_date(soup)
        self.add_company_state(soup)

    @staticmethod
    def page_not_found(browser):
        if browser.current_url.find('ErrorMsg') == -1:
            return False
        else:
            return True

    def get_data_make_table(self, browser, stock_id, code):
        try:
            page = browser.page_source
            soup = BeautifulSoup(page, 'lxml')
            if self.page_not_found(browser):
                return
            hole_data = codal_table.final_table(soup)
            self.save_attribute().insert(0, str(browser.current_url))
            make_file.make_excel(stock_id + '-' + self.sub_id() + '-' + code + '0-' + self.make_name(), stock_id, hole_data,
                       self.save_attribute())
        except:
            make_file.make_error_file(stock_id, code, self.link)

    def give_activity_link_get_table(self, browser, stock_id):
        main_menu, browser = self.open_report_page(self.link, browser)
        print('activity report for link :', self.link)
        for i in range(1, len(main_menu) + 1):
            self.click_menu(browser, i)
            if main_menu[i - 1].get_text().find('گزارش فعالیت ماهانه') != -1:
                self.get_data_make_table(browser, stock_id, '01-')

    def symbol_analysis(self, stock_name, stock_id):
        if self.symbol is None:
            return
        print(stock_name, stock_id, self.symbol, self.link)
        print(self.link)
        sub_id = re.search(r'\d+', self.symbol)
        if sub_id is None:
            make_file.make_sub_file(stock_id, self.company, None, self.code_company_state())
        else:
            make_file.make_sub_file(stock_id, self.company, sub_id.group(), self.code_company_state())


class fiscal(stock_codal):
    def __init__(self):
        self.correction = None
        self.audit = None
        self.company = None
        self.symbol = None
        self.capital = None
        self.unauthorized_capital = None
        self.period = None
        self.term = None
        self.period_end = None
        self.date = None
        self.company_state = None
        self.link = None

    def print_fiscal(self):
        print(self.symbol, self.audit, self.correction, self.company, self.period_end, self.term, self.period \
              , self.capital, self.unauthorized_capital, self.date, self.company_state, self.link)

    def give_link(self, link):
        self.my_link = link

    def print_link(self):
        print(self.my_link)

    def add_correction(self, correction):
        self.correction = correction

    def add_link(self, link):
        self.link = link

    def add_term(self, term):
        self.term = term

    def add_audit(self, audit):
        self.audit = audit

    def add_company(self, soup):
        company_name = soup.find(id=self.company_id)
        if company_name is None:
            return
        company_name = company_name.get_text()
        self.company = str(company_name)

    def add_symbol(self, soup):
        symbol_name = soup.find(id=self.symbol_id)
        if symbol_name is None:
            return
        symbol_name = symbol_name.get_text()
        self.symbol = str(symbol_name)

    def add_capital(self, soup):
        capital_name = soup.find(id=self.capital_id)
        if capital_name is None:
            return
        capital_name = capital_name.get_text()
        capital_name = capital_name.replace(',', '')
        self.capital = int(capital_name)

    def add_unauthorized_capital(self, soup):
        unauthorized_capital_name = soup.find(id=self.unauthorized_capital_id)
        if unauthorized_capital_name is None:
            return
        unauthorized_capital_name = unauthorized_capital_name.get_text()
        unauthorized_capital_name = unauthorized_capital_name.replace(',', '')
        self.unauthorized_capital = str(unauthorized_capital_name)

    def add_period(self, soup):
        period_name = soup.find(id=self.period_id)
        if period_name is None:
            return
        period_name = period_name.get_text()
        period_name = self.find_period(period_name)
        self.period = int(period_name)

    def add_period_end(self, soup):
        period_end_name = soup.find(id=self.period_end_id)
        if period_end_name is None:
            return
        period_end_name = period_end_name.get_text()
        period_end_name = self.find_month(period_end_name)
        self.period_end = str(period_end_name)

    def add_date(self, soup):
        date_name = soup.find(id=self.date_id)
        if date_name is None:
            return
        date_name = date_name.get_text()
        y = date_name.find('13')
        date_name = date_name[y:y + 4]
        self.date = str(date_name)

    def add_company_state(self, soup):
        company_state_name = soup.find(id=self.company_state_id)
        if company_state_name is None:
            return
        company_state_name = company_state_name.get_text()
        company_state_name = self.check_state(company_state_name)
        self.company_state = str(company_state_name)

    def make_name(self):
        name = ''
        if self.audit:
            name += '1-'
        else:
            name += '0-'
        if self.correction:
            name += '1-'
        else:
            name += '0-'
        name += str(self.date) + '-' + str(self.period_end) + '-' + str(self.period)
        return name

    def save_attribute(self):
        att = [self.company, self.symbol, self.capital, self.unauthorized_capital, self.company_state]
        return att

    def code_company_state(self):
        if self.company_state == 'Stock morket':
            return '01'
        elif self.company_state == 'OTC market':
            return '02'
        elif self.company_state == 'Other':
            return '03'
        else:
            return 'N/A'

    def sub_id(self):
        x = re.search(r'\d+', self.symbol)
        if x is None:
            return '00'
        elif int(x.group()) < 10:
            return '0' + x.group()
        elif int(x.group()) >= 10:
            return x.group()

    def check_existence(self):
        if self.company is None:
            return False
        if self.symbol is None:
            return False
        if self.capital is None:
            return False
        return True

    def fiscal_analysis_1(self, line_span, line_td, audit):
        if audit:
            self.add_audit(True)
        else:
            self.add_audit(False)
        if line_span.find('(اصلاحیه)') != -1:
            self.add_correction(True)
        else:
            self.add_correction(False)
        self.add_link(self.find_link(line_td))

    def fiscal_analysis_2(self, soup):
        self.add_company(soup)
        self.add_symbol(soup)
        self.add_capital(soup)
        self.add_unauthorized_capital(soup)
        self.add_period(soup)
        self.add_period_end(soup)
        self.add_date(soup)
        self.add_company_state(soup)

    @staticmethod
    def page_not_found(browser):
        if browser.current_url.find('ErrorMsg') == -1:
            return False
        else:
            return True

    def get_data_make_table(self, browser, stock_id, code, main_menu):
        try:
            page = browser.page_source
            soup = BeautifulSoup(page, 'lxml')
            if self.page_not_found(browser):
                return
            hole_data = codal_table.final_table(soup)
            self.save_attribute().insert(0, browser.current_url)
            if main_menu != -1:
                make_file.make_excel(stock_id + '-' + self.sub_id() + '-' + code + '1-' + self.make_name(), stock_id, hole_data,
                           self.save_attribute())
            else:
                make_file.make_excel(stock_id + '-' + self.sub_id() + '-' + code + '0-' + self.make_name(), stock_id, hole_data,
                           self.save_attribute())
        except:
            make_file.make_error_file(stock_id, code, self.link)

    def give_fiscal_link_get_table(self, browser, stock_id):
        main_menu, browser = self.open_report_page(self.link, browser)
        print('fiscal report for link :', self.link)
        for i in range(1, len(main_menu) + 1):
            self.click_menu(browser, i)
            talfig = main_menu[i - 1].get_text().find('تلفیقی')
            if main_menu[i - 1].get_text().find('ترازنامه') != -1:
                self.get_data_make_table(browser, stock_id, '02-', talfig)
            if main_menu[i - 1].get_text().find('صورت سود و زیان') != -1:
                self.get_data_make_table(browser, stock_id, '03-', talfig)
            if main_menu[i - 1].get_text().find('جریان وجوه نقد') != -1:
                self.get_data_make_table(browser, stock_id, '04-', talfig)

    def symbol_analysis(self, stock_name, stock_id):
        if self.symbol is None:
            return
        sub_id = re.search(r'\d+', self.symbol)
        if sub_id is None:
            make_file.make_sub_file(stock_id, self.company, None, self.code_company_state())
        else:
            make_file.make_sub_file(stock_id, self.company, sub_id.group(), self.code_company_state())


import datetime

begin_time = datetime.datetime.now()


def codal_search_for_links(start, end):
    names = stock_codal.get_stock_names(stock_codal.stock_file_names, stock_codal.id_column, stock_codal.name_column,
                                        start, end)
    browser = stock_codal.open_browser(stock_codal.path)
    bad_stock_names = []

    for k in names:
        make_file.make_folder(k.id)

    for stock in names:
        print(stock.id, stock.name)
        stock.open_first_page(browser)
        li = stock.find_all_li(browser)
        try:
            number_of_pages = stock.find_num_pages(li)
        except:
            print('No data on this page')
            continue
        stock_url = browser.current_url

        for counting in range(1, number_of_pages + 1):
            print('SUCCESS___I could open page', counting, '.')
            page = browser.page_source
            soup = BeautifulSoup(page, 'lxml')
            my_table = soup.find('table')
            t_body = my_table.find("tbody")

            for line_tr in t_body.find_all("tr"):
                fiscal_data = fiscal()
                activity_data = activity()
                line_td = line_tr.find_all("td")[3]
                line_span = line_td.find('span')
                line_span = line_span.get_text()
                line_span = line_span.replace('\n', '')

                if line_span.find('(حسابرسی نشده)') != -1 and \
                        line_span.find('پیش بینی') == -1 and line_span.find('(به پیوست)') == -1:
                    fiscal_data.fiscal_analysis_1(line_span, line_td, False)
                    stock.get_fiscal(fiscal_data)

                elif line_span.find('(حسابرسی شده)') != -1 and \
                        line_span.find('پیش بینی') == -1 and line_span.find('(به پیوست)') == -1:
                    fiscal_data.fiscal_analysis_1(line_span, line_td, True)
                    stock.get_fiscal(fiscal_data)

                elif line_span.find('گزارش فعالیت ماهانه') != -1:
                    activity_data.add_link(activity_data.find_link(line_td))
                    stock.get_activity(activity_data)

            print(browser.current_url)
            browser.get(stock.next_page_address(browser, stock_url, counting))
            time.sleep(2)
            print("SUCCESS___End extracting of page number " + str(counting) + " for stock " + stock.name)
            counting += 1
        print("End finding links of stock", stock.name)
    for stock in names:
        if not stock.status:
            continue
        print('Extraxting tag informations for stock ', stock.name)
        stock.print_num_activities()
        j = 0
        while j < stock.num_activities:
            browser.get(stock.all_activities[j].link)
            time.sleep(2)
            page = browser.page_source
            soup = BeautifulSoup(page, 'lxml')
            stock.all_activities[j].activity_analysis(soup)
            if not stock.all_activities[j].check_existence():
                stock.remove_activity(j)
            else:
                stock.all_activities[j].give_activity_link_get_table(browser, stock.id)
            j += 1
        u = 0
        stock.print_num_fiscals()
        while u < stock.num_fiscals:
            browser.get(stock.all_fiscals[u].link)
            time.sleep(2)
            page = browser.page_source
            soup = BeautifulSoup(page, 'lxml')
            stock.all_fiscals[u].fiscal_analysis_2(soup)
            if not stock.all_fiscals[u].check_existence():
                stock.remove_fiscal(u)
            else:
                stock.all_fiscals[u].give_fiscal_link_get_table(browser, stock.id)
            u += 1
    for stock in names:
        print('Extraxting tables for stock ', stock.name)
        for j in range(stock.num_activities):
            stock.all_activities[j].symbol_analysis(stock.name, stock.id)
            print('activity number ', j + 1, ':')
            stock.all_activities[j].print_activity()
        for u in range(stock.num_fiscals):
            stock.all_fiscals[u].symbol_analysis(stock.name, stock.id)
            print('fiscal number ', u + 1, ':')
            stock.all_fiscals[u].print_fiscal()
    browser.quit()
    print(datetime.datetime.now() - begin_time)
    print('The END')


codal_search_for_links(1,426)


"""
x = 'from codal_main_function import *\ncodal_search_for_links(1,10)\n'
for i in range(1, 426):
    if i < 10:
        start='00'+str(i)
    elif i < 100:
        start = '0'+str(i)
    else:
        start=str(i)
    y=x.replace('(1,10)','('+str(i)+','+str(i)+')')
    f=open('execute_coda_'+start+'.py', mode='w')
    f.write(y)
    f.close()"""
